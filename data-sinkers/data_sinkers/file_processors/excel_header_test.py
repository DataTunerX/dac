"""
Excel header auto-detection regression suite (20 synthetic cases + optional kaogu).

Run:
    PYTHONPATH=. python -m data_sinkers.file_processors.excel_header_test
    PYTHONPATH=. python -m unittest data_sinkers.file_processors.excel_header_test -v
"""

import os
import unittest
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, List, Optional

import pandas as pd
from openpyxl import Workbook

# Heuristic-only in tests (no LLM calls).
os.environ["EXCEL_HEADER_LLM"] = "false"

from .excel import ExcelProcessor
from .excel_header import (
    auto_detect_data_start_row,
    build_column_names,
    skip_leading_empty_rows,
)

FIXTURE_DIR = Path(__file__).resolve().parent / "fixtures" / "excel_headers"
KAOGU_CANDIDATES = [
    os.environ.get("KAOGU_XLSX", ""),
    "/Users/james/daocloud/code/raytest/dac/tests-data/files/kaogu.xlsx",
]


@dataclass
class HeaderCase:
    id: str
    title: str
    description: str
    build: Callable[[str], None]
    expected_header_rows: int
    expected_doc_count: int
    first_doc_must_contain: List[str]
    column_name_must_contain: List[str]


def _write_rows(path: str, rows: List[List]) -> None:
    pd.DataFrame(rows).to_excel(path, index=False, header=False)


def _build_case01_single_row(path: str) -> None:
    _write_rows(path, [["Name", "Age"], ["Tom", 30], ["Jerry", 25]])


def _build_case02_leading_empty(path: str) -> None:
    _write_rows(
        path,
        [["", ""], ["", ""], ["SKU", "Qty"], ["X1", 5], ["X2", 8]],
    )


def _build_case03_two_row_nested(path: str) -> None:
    _write_rows(
        path,
        [
            ["ID", "Group", "", "Stats", ""],
            ["", "Code", "Label", "W", "H"],
            [1, "C1", "Alpha", 10, 20],
            [2, "C2", "Beta", 11, 21],
        ],
    )


def _build_case04_three_row_nested(path: str) -> None:
    _write_rows(
        path,
        [
            ["A", "", "B", ""],
            ["Id", "", "Dim", ""],
            ["No", "Code", "Len", "Wid"],
            [1, "K1", 10, 20],
            [2, "K2", 11, 22],
        ],
    )


def _build_case05_merged_cells(path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.merge_cells("A1:B1")
    ws["A1"] = "Left Block"
    ws.merge_cells("C1:D1")
    ws["C1"] = "Right Block"
    ws["A2"] = "L1"
    ws["B2"] = "L2"
    ws["C2"] = "R1"
    ws["D2"] = "R2"
    ws.append(["a", "b", "c", "d"])
    ws.append(["e", "f", "g", "h"])
    wb.save(path)


def _build_case06_no_serial(path: str) -> None:
    _write_rows(path, [["Person", "Score"], ["Ann", 88], ["Ben", 72]])


def _build_case07_serial_zero(path: str) -> None:
    _write_rows(path, [["N", "V"], [0, "a"], [1, "b"], [2, "c"]])


def _build_case08_title_row(path: str) -> None:
    _write_rows(
        path,
        [
            ["Monthly Performance Overview", "", ""],
            ["Team", "Target", "Actual"],
            ["East", 100, 95],
            ["West", 80, 88],
        ],
    )


def _build_case09_wide_numeric(path: str) -> None:
    _write_rows(
        path,
        [
            ["", "H1", "", "H2", ""],
            ["ID", "M1", "M2", "M3", "M4"],
            [1, 10, 11, 20, 21],
            [2, 12, 13, 22, 23],
        ],
    )


def _build_case10_kaogu_like(path: str) -> None:
    _write_rows(
        path,
        [
            ["", "", "", ""],
            ["No", "Cat", "", "Name"],
            ["", "Type", "Num", ""],
            [1, "R", "1", "Item A"],
            [2, "R", "2", "Item B"],
        ],
    )


def _build_case11_four_row_deep(path: str) -> None:
    """Batch-2: 四层嵌套表头."""
    _write_rows(
        path,
        [
            ["Org", "", "", "Metrics", "", ""],
            ["Unit", "", "Item", "", "", ""],
            ["Field", "Code", "Label", "Len", "Wid", "Hgt"],
            ["No", "", "", "", "", ""],
            [1, "U1", "Part A", 5, 6, 7],
            [2, "U2", "Part B", 8, 9, 10],
        ],
    )


def _build_case12_text_only_data(path: str) -> None:
    """Batch-2: 数据区无数字，全文本."""
    _write_rows(
        path,
        [
            ["Code", "Description"],
            [
                "ART-01",
                "Handmade pottery recovered from the northern sector of the site",
            ],
            [
                "ART-02",
                "Bronze fragment with engraved pattern along the outer rim",
            ],
        ],
    )


def _build_case13_serial_high_start(path: str) -> None:
    """Batch-2: 序号从大数开始递增."""
    _write_rows(
        path,
        [
            ["Idx", "Token"],
            [501, "t501"],
            [502, "t502"],
            [503, "t503"],
        ],
    )


def _build_case14_minimal_one_data(path: str) -> None:
    """Batch-2: 仅 1 行表头 + 1 行数据."""
    _write_rows(path, [["Key", "Val"], ["only", "row"]])


def _build_case15_wide_matrix(path: str) -> None:
    """Batch-2: 12 列宽表 + 双层表头."""
    _write_rows(
        path,
        [
            ["Meta", "", "Q1", "", "Q2", "", "Q3", "", "Q4", "", "Sum", ""],
            ["ID", "Tag", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J"],
            [1, "x", 1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
            [2, "y", 2, 3, 4, 5, 6, 7, 8, 9, 10, 11],
            [3, "z", 3, 4, 5, 6, 7, 8, 9, 10, 11, 12],
        ],
    )


def _build_case16_padded_whitespace(path: str) -> None:
    """Batch-2: 表头单元格含前后空白."""
    _write_rows(
        path,
        [
            ["  Code  ", "  Value  ", "  Note  "],
            ["A1", 42, "ok"],
            ["B2", 17, "fine"],
        ],
    )


def _build_case17_sparse_data_cells(path: str) -> None:
    """Batch-2: 数据行大量空单元格."""
    _write_rows(
        path,
        [
            ["ColA", "ColB", "ColC", "ColD"],
            ["fill", None, 9, None],
            [None, "only", None, 3],
            ["mix", 1, None, None],
        ],
    )


def _build_case18_two_banners(path: str) -> None:
    """Batch-2: 两个标题行 + 表头 + 数据."""
    _write_rows(
        path,
        [
            ["Annual Inventory Summary Report", "", ""],
            ["Warehouse Zone Classification", "", ""],
            ["Slot", "Item", "Count"],
            ["S1", "Bolt", 120],
            ["S2", "Nut", 340],
        ],
    )


def _build_case19_staggered_subheaders(path: str) -> None:
    """Batch-2: 上层稀疏、下层补全子表头."""
    _write_rows(
        path,
        [
            ["Block-A", "", "Block-B", ""],
            ["", "A1", "A2", "B1", "B2"],
            [1, "va", "vb", "wc", "wd"],
            [2, "xa", "xb", "yc", "yd"],
        ],
    )


def _build_case20_negative_and_decimal(path: str) -> None:
    """Batch-2: 负数与小数数据."""
    _write_rows(
        path,
        [
            ["ID", "Delta", "Ratio"],
            [1, -5.5, 0.25],
            [2, -12, 1.75],
            [3, 3.14, -0.5],
        ],
    )


HEADER_CASES: List[HeaderCase] = [
    HeaderCase(
        id="case01_single_row",
        title="单行表头",
        description="标准一行字段名",
        build=_build_case01_single_row,
        expected_header_rows=1,
        expected_doc_count=2,
        first_doc_must_contain=["Name: Tom", "Age: 30"],
        column_name_must_contain=["Name", "Age"],
    ),
    HeaderCase(
        id="case02_leading_empty",
        title="顶部空行",
        description="跳过空白行后识别表头",
        build=_build_case02_leading_empty,
        expected_header_rows=1,
        expected_doc_count=2,
        first_doc_must_contain=["SKU: X1", "Qty: 5"],
        column_name_must_contain=["SKU", "Qty"],
    ),
    HeaderCase(
        id="case03_two_row_nested",
        title="两行嵌套",
        description="分组 + 子字段",
        build=_build_case03_two_row_nested,
        expected_header_rows=2,
        expected_doc_count=2,
        first_doc_must_contain=["Label: Alpha", "Code: C1"],
        column_name_must_contain=["Group", "Label"],
    ),
    HeaderCase(
        id="case04_three_row_nested",
        title="三行嵌套",
        description="三层表头",
        build=_build_case04_three_row_nested,
        expected_header_rows=3,
        expected_doc_count=2,
        first_doc_must_contain=["Code: K1", "Len: 10"],
        column_name_must_contain=["No", "Code"],
    ),
    HeaderCase(
        id="case05_merged_cells",
        title="合并单元格",
        description="openpyxl merge 双区块",
        build=_build_case05_merged_cells,
        expected_header_rows=2,
        expected_doc_count=2,
        first_doc_must_contain=["L1: a", "R2: d"],
        column_name_must_contain=["Left Block", "L1"],
    ),
    HeaderCase(
        id="case06_no_serial",
        title="无序号列",
        description="首列文本",
        build=_build_case06_no_serial,
        expected_header_rows=1,
        expected_doc_count=2,
        first_doc_must_contain=["Person: Ann", "Score: 88"],
        column_name_must_contain=["Person", "Score"],
    ),
    HeaderCase(
        id="case07_serial_zero",
        title="序号从 0",
        description="0,1,2 递增",
        build=_build_case07_serial_zero,
        expected_header_rows=1,
        expected_doc_count=3,
        first_doc_must_contain=["V: a"],
        column_name_must_contain=["N", "V"],
    ),
    HeaderCase(
        id="case08_title_row",
        title="标题行",
        description="长标题 + 表头",
        build=_build_case08_title_row,
        expected_header_rows=2,
        expected_doc_count=2,
        first_doc_must_contain=["Team: East", "Target: 100"],
        column_name_must_contain=["Team", "Target"],
    ),
    HeaderCase(
        id="case09_wide_numeric",
        title="宽表数值",
        description="两层时间分组",
        build=_build_case09_wide_numeric,
        expected_header_rows=2,
        expected_doc_count=2,
        first_doc_must_contain=["M1: 10", "M3: 20"],
        column_name_must_contain=["H1", "M1"],
    ),
    HeaderCase(
        id="case10_kaogu_like",
        title="仿考古台账",
        description="空行 + 双行表头",
        build=_build_case10_kaogu_like,
        expected_header_rows=2,
        expected_doc_count=2,
        first_doc_must_contain=["Name: Item A", "Cat / Type: R"],
        column_name_must_contain=["Cat", "Name"],
    ),
    HeaderCase(
        id="case11_four_row_deep",
        title="四层表头",
        description="Org/Unit/Field/No 四层嵌套",
        build=_build_case11_four_row_deep,
        expected_header_rows=4,
        expected_doc_count=2,
        first_doc_must_contain=["Label: Part A", "Code: U1"],
        column_name_must_contain=["Org", "Label"],
    ),
    HeaderCase(
        id="case12_text_only_data",
        title="纯文本数据",
        description="数据区无数字",
        build=_build_case12_text_only_data,
        expected_header_rows=1,
        expected_doc_count=2,
        first_doc_must_contain=["Code: ART-01", "pottery"],
        column_name_must_contain=["Code", "Description"],
    ),
    HeaderCase(
        id="case13_serial_high_start",
        title="大序号起点",
        description="501-503 连续",
        build=_build_case13_serial_high_start,
        expected_header_rows=1,
        expected_doc_count=3,
        first_doc_must_contain=["Token: t501"],
        column_name_must_contain=["Idx", "Token"],
    ),
    HeaderCase(
        id="case14_minimal_one_data",
        title="最小表",
        description="1 表头 + 1 数据",
        build=_build_case14_minimal_one_data,
        expected_header_rows=1,
        expected_doc_count=1,
        first_doc_must_contain=["Key: only", "Val: row"],
        column_name_must_contain=["Key", "Val"],
    ),
    HeaderCase(
        id="case15_wide_matrix",
        title="12 列宽表",
        description="双层表头 + 多列数据",
        build=_build_case15_wide_matrix,
        expected_header_rows=2,
        expected_doc_count=3,
        first_doc_must_contain=["Tag: x", "A: 1", "J: 10"],
        column_name_must_contain=["Q1", "ID", "Tag"],
    ),
    HeaderCase(
        id="case16_padded_whitespace",
        title="空白填充",
        description="表头 trim 后仍正确",
        build=_build_case16_padded_whitespace,
        expected_header_rows=1,
        expected_doc_count=2,
        first_doc_must_contain=["Code: A1", "Value: 42"],
        column_name_must_contain=["Code", "Value"],
    ),
    HeaderCase(
        id="case17_sparse_data_cells",
        title="稀疏数据行",
        description="数据行多空单元格",
        build=_build_case17_sparse_data_cells,
        expected_header_rows=1,
        expected_doc_count=3,
        first_doc_must_contain=["ColA: fill", "ColC: 9"],
        column_name_must_contain=["ColA", "ColB"],
    ),
    HeaderCase(
        id="case18_two_banners",
        title="双标题行",
        description="两行长标题 + 表头",
        build=_build_case18_two_banners,
        expected_header_rows=3,
        expected_doc_count=2,
        first_doc_must_contain=["Slot: S1", "Item: Bolt"],
        column_name_must_contain=["Slot", "Item"],
    ),
    HeaderCase(
        id="case19_staggered_subheaders",
        title="交错子表头",
        description="上层稀疏下层补全",
        build=_build_case19_staggered_subheaders,
        expected_header_rows=2,
        expected_doc_count=2,
        first_doc_must_contain=["A1: va", "B2: wd"],
        column_name_must_contain=["Block-A", "A1"],
    ),
    HeaderCase(
        id="case20_negative_and_decimal",
        title="负数小数",
        description="负值与小数数据",
        build=_build_case20_negative_and_decimal,
        expected_header_rows=1,
        expected_doc_count=3,
        first_doc_must_contain=["Delta: -5.5", "Ratio: 0.25"],
        column_name_must_contain=["Delta", "Ratio"],
    ),
]


def _detect_header_rows(file_path: str) -> int:
    raw = pd.read_excel(file_path, header=None)
    raw = skip_leading_empty_rows(raw)
    return auto_detect_data_start_row(raw)


def _detect_column_names(file_path: str) -> List[str]:
    raw = pd.read_excel(file_path, header=None)
    raw = skip_leading_empty_rows(raw)
    data_start = auto_detect_data_start_row(raw)
    if data_start <= 0 or data_start >= len(raw):
        return []
    return build_column_names(raw.iloc[:data_start])


def materialize_fixtures() -> Path:
    FIXTURE_DIR.mkdir(parents=True, exist_ok=True)
    for case in HEADER_CASES:
        case.build(str(FIXTURE_DIR / f"{case.id}.xlsx"))
    return FIXTURE_DIR


def _resolve_kaogu_path() -> Optional[str]:
    for candidate in KAOGU_CANDIDATES:
        if candidate and os.path.isfile(candidate):
            return candidate
    return None


class ExcelHeaderDetectionTest(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.fixture_dir = materialize_fixtures()
        cls.processor = ExcelProcessor()

    def _run_case(self, case: HeaderCase) -> None:
        path = self.fixture_dir / f"{case.id}.xlsx"
        self.assertTrue(path.is_file(), f"fixture missing: {path}")

        detected = _detect_header_rows(str(path))
        self.assertEqual(
            detected,
            case.expected_header_rows,
            f"{case.id}: expected {case.expected_header_rows} header rows, got {detected}",
        )

        docs = self.processor.process_excel(str(path))
        self.assertEqual(
            len(docs),
            case.expected_doc_count,
            f"{case.id}: expected {case.expected_doc_count} documents, got {len(docs)}",
        )

        content = docs[0].page_content
        for snippet in case.first_doc_must_contain:
            self.assertIn(snippet, content, f"{case.id}: missing '{snippet}' in first doc")

        columns = _detect_column_names(str(path))
        joined = " | ".join(columns)
        for snippet in case.column_name_must_contain:
            self.assertIn(snippet, joined, f"{case.id}: columns missing '{snippet}' in {columns}")

    def test_batch1_cases(self) -> None:
        for case in HEADER_CASES[:10]:
            with self.subTest(case=case.id):
                self._run_case(case)

    def test_batch2_cases(self) -> None:
        for case in HEADER_CASES[10:]:
            with self.subTest(case=case.id):
                self._run_case(case)

    def test_real_kaogu_if_available(self) -> None:
        kaogu_path = _resolve_kaogu_path()
        if not kaogu_path:
            self.skipTest("kaogu.xlsx not found")

        detected = _detect_header_rows(kaogu_path)
        self.assertEqual(detected, 2)

        docs = self.processor.process_excel(kaogu_path)
        self.assertEqual(len(docs), 465)

        sample = next(d for d in docs if "五铢" in d.page_content)
        self.assertIn("*名", sample.page_content)


def print_case_report(batch: Optional[str] = None) -> None:
    fixture_dir = materialize_fixtures()
    processor = ExcelProcessor()
    cases = HEADER_CASES
    if batch == "1":
        cases = HEADER_CASES[:10]
    elif batch == "2":
        cases = HEADER_CASES[10:]

    print(f"Excel header detection report ({len(cases)} cases)")
    print("=" * 100)
    passed = 0
    for case in cases:
        path = fixture_dir / f"{case.id}.xlsx"
        detected = _detect_header_rows(str(path))
        docs = processor.process_excel(str(path))
        columns = _detect_column_names(str(path))
        ok = detected == case.expected_header_rows and len(docs) == case.expected_doc_count
        passed += int(ok)
        status = "PASS" if ok else "FAIL"
        print(f"\n[{status}] {case.id} — {case.title}")
        print(f"  {case.description}")
        print(f"  header: expect {case.expected_header_rows}, got {detected}")
        print(f"  rows:   expect {case.expected_doc_count}, got {len(docs)}")
        for col_idx, col_name in enumerate(columns):
            print(f"    C{col_idx}: {col_name}")

    print(f"\n{passed}/{len(cases)} passed")
    print(f"Fixtures: {fixture_dir}")


if __name__ == "__main__":
    print_case_report()
    unittest.main(argv=[""], verbosity=2, exit=False)
